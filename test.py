import requests
import openpyxl
from openpyxl import Workbook

def validate_email(email):
    print(f"Validating email {email}...")
    api_key = "at_sm3Xy5aowsOCoyDHv5oz2gVcjBNL5"  # replace with your actual API key
    api_url = "null" # replace with your actual API URL

    try:
        response = requests.get(api_url)
        response.raise_for_status()  # Raises stored HTTPError, if one occurred.

    except requests.exceptions.HTTPError as errh:
        # Check for different error status codes
        if response.status_code == 400:
            print(f"Error: Bad Request for email {email}. Details: {errh}")
        elif response.status_code == 401:
            print(f"Error: Unauthorized for email {email}. Details: {errh}")
        elif response.status_code == 402:
            print(f"Error: Payment Required for email {email}. Details: {errh}")
        elif response.status_code == 403:
            print(f"Error: Forbidden for email {email}. Details: {errh}")
        elif response.status_code == 422:
            print(f"Error: Unprocessable Entity for email {email}. Details: {errh}")
        elif response.status_code == 429:
            print(f"Error: Too Many Requests for email {email}. Details: {errh}")
        else:
            print(f"Error: An HTTP error occurred for email {email}. Details: {errh}")
        return False

    except requests.exceptions.ConnectionError as errc:
        print(f"Error: A Network problem occurred with email {email}. Details: {errc}")
        return False

    except requests.exceptions.Timeout as errt:
        print(f"Error: Request timed out for email {email}. Details: {errt}")
        return False

    except requests.exceptions.RequestException as err:
        print(f"Error: Something went wrong with the request for email {email}. Details: {err}")
        return False

    # Check the smtpCheck field in the response
    data = response.json()
    if data.get('smtpCheck') == 'true':
        return True
    else:
        return False


def process_emails(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    valid_emails = []

    for row in sheet.iter_rows(values_only=True):
        for email in row:
            if validate_email(email):
                valid_emails.append(email)

    # Write valid emails to a new Excel file
    wb_valid = Workbook()
    sheet_valid = wb_valid.active
    for email in valid_emails:
        sheet_valid.append([email])

    wb_valid.save('valid_emails.xlsx')


# use the function with the name of your Excel file
process_emails('emails.xlsx')
