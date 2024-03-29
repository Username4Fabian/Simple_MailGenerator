#!/usr/bin/env python3s

# python -m PyInstaller main.spec

import pandas as pd
import openpyxl
import os
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QLineEdit, QLabel, QMessageBox, QGridLayout, QSizePolicy
from PyQt5.QtGui import QPalette, QColor, QFont
from PyQt5.QtCore import Qt
import json
import requests
import concurrent.futures
import requests
import requests
import concurrent.futures

def main():
    # Read or create email settings
    if os.path.exists("email_settings.txt"):
        with open("email_settings.txt", "r") as f:
            email_settings = json.load(f)
            email_domains = email_settings['email_domains']
            email_format_structures = email_settings['email_format_structures']
            api_key = email_settings['api_key']  # add this line


        # generate email_formats
        email_formats = []
        for x in email_domains:
            email_formats += [format + x for format in email_format_structures]
    else:
        # If no settings file found, use these default values and write them to a new file
        email_domains =  ["@gmail.com", "@aon.at", "@gmx.at", "@gmx.net", "@outlook.com", "@icloud.com"]
        email_format_structures = ["{f}.{l}", "{f}{l}", "{f}_{l}", "{f[0]}.{l}", "{f}.{l[0]}", "{l}{f}"]
        api_key = "at_sm3Xy5aowsOCoyDHv5oz2gVcjBNL5" #API Key for email verification


        # generate email_formats
        email_formats = []
        for x in email_domains:
            email_formats += [format + x for format in email_format_structures]

        email_settings = {'email_domains': email_domains, 'email_format_structures': email_format_structures, 'api_key': api_key}

        with open("email_settings.txt", "w") as f:
            json.dump(email_settings, f)

    if(os.path.exists("emails.xlsx") == False):
        filename = 'emails.xlsx'
        workbook = openpyxl.Workbook()
        workbook.save(filename)


    # Email presets for first name only
    first_name_email_formats = ["{f}"+x for x in email_domains]

    # Function to generate email addresses
    def generate_emails(first_name, last_name):
        emails = []
        if last_name:
            for email_format in email_formats:
                email = email_format.format(f=first_name, l=last_name)
                if validate_email(email):  # Add this check before appending
                    emails.append(email)
                    
        else:
            for email_format in first_name_email_formats:
                email = email_format.format(f=first_name)
                if validate_email(email):  # Add this check before appending
                    emails.append(email)
        return emails
    
    # Add a new function to validate emails
    def validate_email(email):
        def fetch_url():
            print(f"Validating email {email}...")
            api_url = f"https://emailverification.whoisxmlapi.com/api/v2?apiKey={api_key}&emailAddress={email}"
            return requests.get(api_url, timeout=3)  # timeout after 3 seconds

        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(fetch_url)
            try:
                response = future.result()
                response.raise_for_status()  # raise an exception for non-200 status codes
                data = response.json()
                if data.get('smtpCheck') == 'true':
                    print("Valid email")
                    return True
                else:
                    print("Invalid email")
                    return False
            except requests.exceptions.Timeout:
                print("Timeout, not a valid email")
                return False
            except requests.exceptions.RequestException as e:
                print(f"Error occurred: {e}")
                return False
    # Function to get names from user
    def get_names():
        names = []
        line_counter = 1  # Initialize line counter

        full_name = name_entry.text().strip().lower()
        if full_name == '':
            QMessageBox.critical(window, "Error", "Please enter a name")
            return

        name_parts = full_name.split(" ")
        if len(name_parts) == 2:
            first_name, last_name = name_parts
        else:
            first_name = name_parts[0]
            last_name = ""

        names.append((first_name, last_name))
        
        # Load the workbook
        workbook = openpyxl.load_workbook('emails.xlsx')

        # Select the first sheet (you can modify this if you have multiple sheets)
        sheet = workbook.active

        # Find the last line counted in column A
        while sheet.cell(row=line_counter, column=1).value is not None:
            line_counter += 1

        # Generate emails for the current name
        all_emails = generate_emails(first_name, last_name)

        # Write the emails to the next line
        for email in all_emails:
            sheet.cell(row=line_counter, column=1).value = email

            # Increment the line counter
            line_counter += 1

        # Save the workbook
        workbook.save('emails.xlsx')


        name_entry.clear() # clear the input field
        QMessageBox.information(window, "Success", f"Emails for {full_name} have been generated")

    # New function to open the Excel file
    def open_excel():
        filename = 'emails.xlsx'
        if os.path.isfile(filename):
            if os.name == 'nt':
                os.system('start excel.exe "%s"' % filename)
            elif os.name == 'posix':
                os.system('open "%s"' % filename)
            else:
                QMessageBox.critical(window, "Error", "OS not supported")
        else:
            QMessageBox.critical(window, "Error", "File not found")


    # GUI setup
    app = QApplication([])
    window = QWidget()

    # Set the window size and background color
    window.resize(1200, 600)  # Set window size to 800x600
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor(44, 47, 51))  # Discord-like dark background color
    window.setPalette(palette)

    outer_layout = QVBoxLayout()  # Outer layout to center the grid layout vertically
    layout = QGridLayout()  # Grid layout to center widgets horizontally

    header_label = QLabel("E-Mail Generator")
    header_label.setFont(QFont('Arial', 30))
    header_label.setStyleSheet("color: white")
    header_label.setAlignment(Qt.AlignCenter)  # Align the header to the center
    header_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    name_label = QLabel("Full Name:")
    name_label.setFont(QFont('Arial', 22))  # Make the "Full Name:" label a bit smaller
    name_label.setStyleSheet("color: white")  # Set label text color to white like Discord
    name_label.setAlignment(Qt.AlignCenter)  # Align the label to the center
    name_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    name_entry = QLineEdit()
    name_entry.setStyleSheet("background-color: white; color: black")  # Set entry field background to white and text to black
    name_entry.setFont(QFont('Arial', 18))
    name_entry.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
    generate_button = QPushButton("Generate Emails")
    generate_button.setStyleSheet("background-color: #7289DA; color: white; font-size: 22px")  # Discord-like button color
    generate_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Make the button expand to fill available space
    generate_button.clicked.connect(get_names)
    name_entry.returnPressed.connect(get_names)

    # New button to open the Excel file
    open_button = QPushButton("Open Excel File")
    open_button.setStyleSheet("background-color: #7289DA; color: white; font-size: 22px")  # Discord-like button color
    open_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding) 
    open_button.clicked.connect(open_excel)

    # Place the widgets in the middle of the grid
    layout.addWidget(header_label, 0, 0)
    layout.addWidget(name_label, 1, 0)
    layout.addWidget(name_entry, 2, 0)
    layout.addWidget(generate_button, 3, 0)
    layout.addWidget(open_button, 4, 0)  # Add the new button to the grid layout
    layout.setVerticalSpacing(20)  

    # Add the grid layout to the outer layout
    outer_layout.addStretch()
    outer_layout.addLayout(layout)
    outer_layout.addStretch()

    window.setLayout(outer_layout)
    window.show()

    app.exec_()


if __name__ == "__main__":
    main()